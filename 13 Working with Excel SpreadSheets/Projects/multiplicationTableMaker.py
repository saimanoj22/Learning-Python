# usage: py multiplicationTableMaker.py 6
import openpyxl, sys
from openpyxl.styles import Font

# Command line input
n = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'hello'
fontObj = Font(bold=True)
for rowNo in range(1, n+2):
    for columnNo in range(1, n+2):
        if rowNo == 1:
            sheet.cell(row = rowNo, column=columnNo).font = fontObj
            sheet.cell(row = rowNo, column = columnNo).value = columnNo - 1
        elif columnNo == 1:
            sheet.cell(row = rowNo, column=columnNo).font = fontObj
            sheet.cell(row = rowNo, column = columnNo).value = rowNo - 1
        else:
            sheet.cell(row = rowNo, column = columnNo).value = (rowNo - 1) * (columnNo - 1)

del sheet['A1']
wb.save('multiplicationTableMaker.xlsx')
