# usage: python blankRowInserter.py 3 2 myFile.xlsx
import openpyxl, sys

start, noOfBlanks, file = int(sys.argv[1]), int(sys.argv[2]), sys.argv[3]

wb = openpyxl.load_workbook(file)
sheet = wb.active
rowsList = tuple(sheet.rows)

for rowObj in  rowsList[::-1]:
    for cellObj in rowObj:
        col = cellObj.column
        ro = cellObj.row

        if ro >= start and ro < start+noOfBlanks:
            sheet.cell(row=ro+noOfBlanks, column=col).value = cellObj.value
            sheet.cell(row=ro, column=col).value = ''
        elif ro >= start+noOfBlanks:
            sheet.cell(row=ro+noOfBlanks,column=col).value = cellObj.value
file = file.replace('.xlsx','')
wb.save(file+'_output.xlsx')