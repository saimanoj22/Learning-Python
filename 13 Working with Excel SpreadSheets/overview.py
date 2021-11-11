# module for manipulating excel files
from os import name, set_inheritable
import openpyxl

# just like fileObject.open()
wb = openpyxl.load_workbook('example.xslx')
type(wb)    # <class 'openpyxl.workbook.workbook.Workbook'>

# getting list of sheets
wb.sheetnames   # ['Sheet1', 'Sheet2', 'Sheet3']

# getting particular sheet
sheet = wb['Sheet3']    # Get a sheet from the workbook
sheet           # <Worksheet "Sheet3">
type(sheet)     # <class 'openpyxl.worksheet.worksheet.Worksheet'>

# get sheet title as string
sheet.title     # ' Sheet3'

# get active sheet
anotherSheet = wb.active
anotherSheet    # <Worksheet "Sheet1">

# GETTING CELLS FROM THE SHEETS
sheet['A1']         # get a cell from the sheet
# <Cell 'Sheet1'.A1>
sheet['A1'].value   # get the value from the cell
# datetime.datetime(2015, 4, 5, 13, 34, 2)
c = sheet['B1']
c.value
# 'Apples'

# Get the row, column, and value from the cell
'Row %s, Column %s is %s' %(c.row, c.column, c.value)
# 'Row 1, Column B is Apples'
'Cell %s is %s' %(c.coordinate, c.value)
# 'Cell B1 is Apples'

# cell() method
sheet.cell(row=1, column=2)
# <Cell 'Sheet1'.B1>
sheet.cell(row=1,column=2).value
# 'Apples'

# Get highest row and column number
sheet.max_row
sheet.max_column

# Converting between columns letters and numbers
from openpyxl.utils import get_column_letter, column_index_from_string
get_column_letter(1)                    # 'A'
get_column_letter(27)                   # 'AA'
get_column_letter(sheet.max_column)     # 'C'
column_index_from_string('A')           # 1
column_index_from_string('AA')          # 27

# Getting rows and columns from the sheets
tuple(sheet['A1':'C3'])
'''
((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), (<Cell
'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>), (<Cell 'Sheet1'.A3>,
<Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))
'''

list(sheet.columns)[1]      # Get second column's cells


# Renaming active sheet
sheet.active = 'NEW NAME'
# the spreadsheet will not be saved until we call save() workbook method.
wb.save('example_copy.xlsx')

# creating and Removing sheets
wb = openpyxl.Workbook()
wb.sheetnames                                       # ['Sheets']
wb.create_sheet()
wb.sheetnames                                       # ['Sheet','Sheet1']
wb.create_sheet(index=0, title='First Sheet')       # ['First Sheet','Sheet','Sheet1']
wb.create_sheet(index=2, title='Middle Sheet')      # ['First Sheet','Sheet','Middle Sheet','Sheet1']
# Deleting spreadsheets
del wb['Middle Sheet']


# Writing values to cell
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hello, World!'
sheet['A1'].value       # 'Hello, World!'

# Setting the Font Styles of Cells
from openpyxl.styles import Font
italic24Font = Font(size=24, italic=True)   #Create a Font
sheet['A1'].font = italic24Font
sheet['A1'] = 'Hello, world!'

# Font Objects
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb['Sheet']

fontObj1 = Font(name='Times New Roman', bold=True)
sheet['A1'].font = fontObj1
sheet['A1'] = 'Bold Times New Roman'

# Formulas
sheet['B9'] = '=SUM(B1:B8)'

# Adjusting Rows and Columns
sheet['A1'] = 'Tall Row'
sheet['B2'] = 'Wide Column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'] = 20

# Merging and Unmerging Cells
sheet.merge_cells('A1:D3')      # merge all these cells
sheet['A1'] = 'Twelve cells merged together'

sheet.unmerge_cells('A1:D3')    # split these cells up

# Freezing panes
sheet.freeze_panes = 'A2'   # row 1 will br frozen
sheet.freeze_panes = 'B1'   # column a will be frozen

# Charts
for i in range(1, 11):      # create some data in column A
    sheet['A' + str(i)] = i
refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
seriesObj = openpyxl.chart.Series(refObj, title='First series')
chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'C5')

